#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""FP3級 重要用語一覧（xlsx）を glossary_terms.csv に draft 行として追加する."""

from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

GLOSSARY_CSV = ROOT / "data" / "glossary_terms.csv"
DEFAULT_XLSX = Path.home() / "Downloads" / "FP3級_重要用語一覧.xlsx"
SHEET_NAME = "FP3級用語一覧"


def norm(value: object) -> str:
    return str(value or "").strip()


def load_fieldnames() -> list[str]:
    with GLOSSARY_CSV.open(encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            raise ValueError(f"{GLOSSARY_CSV} にヘッダがありません")
        fields = list(reader.fieldnames)
    if "content_status" not in fields:
        fields.append("content_status")
    return fields


def read_existing_rows() -> tuple[list[str], list[dict[str, str]]]:
    fields = load_fieldnames()
    with GLOSSARY_CSV.open(encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f))
    for row in rows:
        if not norm(row.get("content_status")):
            row["content_status"] = "published"
    return fields, rows


def read_xlsx_terms(path: Path) -> list[tuple[str, str, str]]:
    try:
        import openpyxl
    except ImportError as e:
        raise SystemExit("openpyxl が必要です: pip install openpyxl") from e

    if not path.is_file():
        raise FileNotFoundError(path)

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"シート {SHEET_NAME!r} がありません: {wb.sheetnames}")
    ws = wb[SHEET_NAME]
    out: list[tuple[str, str, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        term = norm(row[0])
        category = norm(row[1])
        definition = norm(row[2])
        if not term or not category or not definition:
            raise ValueError(f"用語・科目・定義のいずれかが空です: {row!r}")
        out.append((term, category, definition))
    return out


def build_draft_row(
    term: str,
    category: str,
    definition: str,
    *,
    fieldnames: list[str],
) -> dict[str, str]:
    row = {col: "" for col in fieldnames}
    row.update(
        {
            "term": term,
            "category": category,
            "tags": "FP3級;用語",
            "short_def": definition,
            "definition": definition,
            "importance": "B",
            "content_status": "draft",
        }
    )
    return row


def write_csv(fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    with GLOSSARY_CSV.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--xlsx",
        type=Path,
        default=DEFAULT_XLSX,
        help=f"入力 Excel（既定: {DEFAULT_XLSX}）",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="CSV を書き込まず追加件数のみ表示",
    )
    args = parser.parse_args()

    fieldnames, existing = read_existing_rows()
    existing_terms = {norm(r.get("term")) for r in existing if norm(r.get("term"))}

    xlsx_rows = read_xlsx_terms(args.xlsx)
    to_add: list[dict[str, str]] = []
    skipped = 0
    for term, category, definition in xlsx_rows:
        if term in existing_terms:
            skipped += 1
            continue
        to_add.append(build_draft_row(term, category, definition, fieldnames=fieldnames))
        existing_terms.add(term)

    print(f"Excel: {len(xlsx_rows)} 語")
    print(f"既存 CSV: {len(existing)} 語")
    print(f"追加予定: {len(to_add)} 語（スキップ {skipped} 語）")

    if args.dry_run:
        return 0

    if not to_add:
        print("追加する用語はありません。")
        return 0

    write_csv(fieldnames, existing + to_add)
    print(f"Wrote {len(existing) + len(to_add)} rows to {GLOSSARY_CSV}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
